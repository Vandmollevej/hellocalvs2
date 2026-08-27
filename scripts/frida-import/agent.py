"""
HELLO CAL Frida-agent.

Poller DTU Fødevareinstituttets officielle datarepository (data.dtu.dk, en
Figshare-instans) for en ny udgivelse af Frida (Den Danske Fødevaredatabase),
og importerer den automatisk, når der kommer en ny version — uden manuel
download. Datasættet er offentligt og CC-BY 4.0-licenseret; DTU Foods
Figshare-gruppe har group_id 18053.

Hver fødevare upsertes som et produkt (externalSource=FRIDA, status=APPROVED,
ingen stregkode) matchet på (externalSource, externalId=FoodID), så
genkørsel/en ny version opdaterer eksisterende rækker i stedet for at
duplikere dem. `frida_import_state` husker hvilken Figshare-artikel-id der
senest er importeret, så samme version ikke hentes/importeres igen.

Kildeangivelse (Frida-vilkår): "Fødevaredata (frida.fooddata.dk), DTU
Fødevareinstituttet, Danmarks Tekniske Universitet".
"""

import io
import logging
import os
import time

import openpyxl
import psycopg2
import requests

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("frida-agent")

# Prisma's DATABASE_URL carries a `?schema=public` query param that
# psycopg2/libpq doesn't recognize; Postgres already defaults to "public".
DATABASE_URL = os.environ["DATABASE_URL"].split("?")[0]
POLL_INTERVAL_SECONDS = int(os.environ.get("FRIDA_POLL_INTERVAL_SECONDS", str(24 * 60 * 60)))

FIGSHARE_SEARCH_URL = "https://api.figshare.com/v2/articles/search"
FIGSHARE_GROUP_ID = 18053
TITLE_MATCH = "Danish Food Composition Database"

# Frida ParameterID for hver af HELLO CALs fire kernenæringsstoffer pr. 100 g.
PARAM_KCAL = 356
PARAM_PROTEIN = 218
PARAM_CARBS = 170  # "Kulhydrat difference" — svarer til USDA's "Carbohydrate, by difference".
PARAM_FAT = 141
WANTED_PARAMS = {PARAM_KCAL, PARAM_PROTEIN, PARAM_CARBS, PARAM_FAT}


def find_latest_article():
    response = requests.post(
        FIGSHARE_SEARCH_URL,
        json={
            "search_for": f'"{TITLE_MATCH}"',
            "order": "published_date",
            "order_direction": "desc",
            "page_size": 20,
        },
        timeout=30,
    )
    response.raise_for_status()
    candidates = [
        item
        for item in response.json()
        if item.get("group_id") == FIGSHARE_GROUP_ID and TITLE_MATCH in item.get("title", "")
    ]
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: item["published_date"], reverse=True)[0]


def find_dataset_download_url(article_id):
    response = requests.get(f"https://api.figshare.com/v2/articles/{article_id}", timeout=30)
    response.raise_for_status()
    for file in response.json().get("files", []):
        if file.get("name", "").lower().endswith(".xlsx"):
            return file["download_url"]
    return None


def download_workbook(download_url):
    response = requests.get(download_url, timeout=120)
    response.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)


def parse_foods(workbook):
    names = {}
    for name_dk, name_en, food_id, *_ in workbook["Food"].iter_rows(min_row=2, values_only=True):
        if food_id is not None:
            names[food_id] = (name_dk, name_en)

    values = {}
    for row in workbook["Data_Normalised"].iter_rows(min_row=2, values_only=True):
        food_id, param_id, res_val = row[0], row[3], row[7]
        if param_id in WANTED_PARAMS and isinstance(res_val, (int, float)):
            values.setdefault(food_id, {})[param_id] = float(res_val)

    foods = []
    for food_id, params in values.items():
        if not WANTED_PARAMS.issubset(params):
            continue
        name_dk, _name_en = names.get(food_id, (None, None))
        if not name_dk:
            continue
        foods.append(
            {
                "external_id": str(food_id),
                "name": name_dk.strip(),
                "kcal": round(params[PARAM_KCAL], 2),
                "protein": round(params[PARAM_PROTEIN], 2),
                "carbs": round(params[PARAM_CARBS], 2),
                "fat": round(params[PARAM_FAT], 2),
            }
        )
    return foods


def already_imported(conn, article_id):
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM frida_import_state WHERE \"figshareArticleId\" = %s", (article_id,))
        return cur.fetchone() is not None


def upsert_foods(conn, foods):
    inserted = updated = 0
    with conn.cursor() as cur:
        for food in foods:
            cur.execute(
                """SELECT id FROM products WHERE "externalSource" = 'FRIDA' AND "externalId" = %s""",
                (food["external_id"],),
            )
            existing = cur.fetchone()

            if existing:
                cur.execute(
                    """
                    UPDATE products
                    SET name = %s, "kcalPer100g" = %s, "proteinPer100g" = %s,
                        "carbsPer100g" = %s, "fatPer100g" = %s, "sourceCheckedAt" = NOW()
                    WHERE id = %s
                    """,
                    (food["name"], food["kcal"], food["protein"], food["carbs"], food["fat"], existing[0]),
                )
                updated += 1
            else:
                cur.execute(
                    """
                    INSERT INTO products
                        (id, name, "kcalPer100g", "proteinPer100g", "carbsPer100g", "fatPer100g",
                         "externalSource", "externalId", "sourceCheckedAt", status, discontinued, "createdAt")
                    VALUES
                        (%s, %s, %s, %s, %s, %s, 'FRIDA', %s, NOW(), 'APPROVED', false, NOW())
                    """,
                    (
                        f"frida_{food['external_id']}",
                        food["name"],
                        food["kcal"],
                        food["protein"],
                        food["carbs"],
                        food["fat"],
                        food["external_id"],
                    ),
                )
                inserted += 1
    return inserted, updated


def mark_imported(conn, article_id, title):
    with conn.cursor() as cur:
        cur.execute(
            """INSERT INTO frida_import_state ("figshareArticleId", title) VALUES (%s, %s)""",
            (article_id, title),
        )


def run_once(conn):
    article = find_latest_article()
    if not article:
        log.warning("no Frida dataset found via Figshare search")
        return

    article_id, title = article["id"], article["title"]
    if already_imported(conn, article_id):
        log.info("already imported: %s (%s)", title, article_id)
        return

    log.info("new Frida release found: %s (%s)", title, article_id)
    download_url = find_dataset_download_url(article_id)
    if not download_url:
        log.error("no .xlsx file found on article %s", article_id)
        return

    workbook = download_workbook(download_url)
    foods = parse_foods(workbook)
    log.info("parsed %d foods with all four macros", len(foods))

    inserted, updated = upsert_foods(conn, foods)
    mark_imported(conn, article_id, title)
    conn.commit()
    log.info("import complete — inserted: %d, updated: %d", inserted, updated)


def main():
    log.info("frida agent started, polling every %ss", POLL_INTERVAL_SECONDS)
    while True:
        try:
            with psycopg2.connect(DATABASE_URL) as conn:
                run_once(conn)
        except Exception:  # noqa: BLE001 - a broken cycle must not kill the service
            log.exception("cycle failed")
        time.sleep(POLL_INTERVAL_SECONDS)


if __name__ == "__main__":
    main()
